# excel/io.py
"""
Excel I/O for the product-check workflow.

Spaltenlayout (Header in Zeile 3, Daten ab Zeile 4):
  - ID                       (Input, Pflicht)
  - Available                (Output ja/nein)
  - Min. 6 Bilder            (Output ja/nein)
  - 1- 3 Videos              (Output ja/nein)
  - Produktbeschr            (Output ja/nein)
  - Gutes Mainbild           (Output ja/nein)
  - Keyword                  (Input – Suchwort)
  - Platzierung Keyword      (Output Zahl 1..48 oder leer)
  - Preis Galaxus            (Output Zahl)
  - Preis Marktpreis Main    (Output Zahl)
  - Marktpreis Anbieter Main (Output Text, z.B. "Fust")
"""
from __future__ import annotations

import io
from dataclasses import dataclass
from typing import Dict, List

import openpyxl

HEADER_ROW = 3
DATA_START_ROW = 4


def normalize_header(value) -> str:
    if value is None:
        return ""
    return " ".join(str(value).split())


@dataclass
class InputRow:
    row_index: int
    product_id: str
    keyword: str


def load_workbook_from_bytes(xlsx_bytes: bytes):
    return openpyxl.load_workbook(io.BytesIO(xlsx_bytes))


def workbook_to_bytes(wb) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def map_headers(ws) -> Dict[str, int]:
    headers: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        h = normalize_header(ws.cell(HEADER_ROW, col).value)
        if h:
            headers[h] = col
    return headers


def read_inputs(ws, headers: Dict[str, int]) -> List[InputRow]:
    if "ID" not in headers:
        raise ValueError("Pflicht-Spalte 'ID' fehlt in Zeile 3.")

    id_col = headers["ID"]
    keyword_col = headers.get("Keyword")  # optional

    rows: List[InputRow] = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        pid = normalize_header(ws.cell(r, id_col).value)
        if not pid:
            continue
        kw = normalize_header(ws.cell(r, keyword_col).value) if keyword_col else ""
        rows.append(InputRow(row_index=r, product_id=pid, keyword=kw))
    return rows


def _set(ws, row: int, col: int, value):
    if col:
        ws.cell(row, col).value = value


def write_outputs(ws, headers: Dict[str, int], results_by_row: Dict[int, dict]) -> None:
    """
    Schreibt die Ergebnisse zurück ins Worksheet.
    Erwartete Keys pro Zeile:
      available, images_ok, videos_ok, bullets_ok, mainbild_ok,
      keyword_rank, galaxus_price, toppreise_price, toppreise_vendor
    """
    col_available  = headers.get("Available")
    col_img        = headers.get("Min. 6 Bilder")
    col_vid        = headers.get("1- 3 Videos")
    col_bul        = headers.get("Produktbeschr")
    col_main       = headers.get("Gutes Mainbild")
    col_kw_rank    = headers.get("Platzierung Keyword")
    col_price_gx   = headers.get("Preis Galaxus")
    col_price_mp   = headers.get("Preis Marktpreis Main")
    col_vendor_mp  = headers.get("Marktpreis Anbieter Main")

    for r, res in results_by_row.items():
        _set(ws, r, col_available, res.get("available", "nein"))
        _set(ws, r, col_img,       res.get("images_ok", "nein"))
        _set(ws, r, col_vid,       res.get("videos_ok", "nein"))
        _set(ws, r, col_bul,       res.get("bullets_ok", "nein"))
        _set(ws, r, col_main,      res.get("mainbild_ok", "nein"))
        _set(ws, r, col_kw_rank,   res.get("keyword_rank", ""))
        _set(ws, r, col_price_gx,  res.get("galaxus_price", ""))
        _set(ws, r, col_price_mp,  res.get("toppreise_price", ""))
        _set(ws, r, col_vendor_mp, res.get("toppreise_vendor", ""))
